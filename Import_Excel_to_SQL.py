import os

import openpyxl
#import xlrd
import mysql.connector
from mysql.connector import Error


def create_connection(host_name, user_name, user_password, database_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            password=user_password,
            database=database_name
        )

        print('Connection to MySQL DB successful')
    except Error as e:
        print(f"The error '{e} occurred")

    return connection


def create_database(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        print("Database created successfully")
    except Error as e:
        print(f"The error '{e}' occurred")


def execute_query(connection, query):
    cursor = connection.cursor()

    try:
        cursor.execute(query)
        connection.commit()
        print("Query executed successfully")
    except Error as e:
        print(f"The error '{e}' occurred")


cwd = os.getcwd()

os.chdir("/home/ubuntu/Desktop/Diplom/ветер_факт")

os.listdir(".")

path = "/home/ubuntu/Desktop/Diplom/ветер_факт/USDD_2015_01 (1).xlsx"

book = openpyxl.load_workbook(path)
sheet = book.active

connection = create_connection("localhost", "root", "z1k1v2aA*", "")
create_database_query = "CREATE DATABASE 'USDD_2015_01'"

create_database(connection, create_database_query)
connection = create_connection("localhost", "root", "z1k1v2aA*", "USDD_2015_01")

create_users_table = """
use USDD_2015_01;
CREATE TABLE IF NOT EXISTS USDD_2015_01 (
  Year INT, 
  Month INT, 
  Day INT, 
  Hour INT, 
  Minute INT, 
  WD INT,
  WS INT,
  WSm INT
) ENGINE = InnoDB
"""

execute_query(connection, create_users_table)

cursor = connection.cursor()

create_users = """
INSERT INTO
  `USDD_2015_01` (Year, Month, Day, Hour, Minute, WD, WS, WSm)
VALUES
                 (%s, %s, %s, %s, %s, %s, %s, %s)
"""


for r in range(1, sheet.max_row):
    Year = sheet.cell(r, 1).value
    Month = sheet.cell(r, 2).value
    Day = sheet.cell(r, 3).value
    Hour = sheet.cell(r, 4).value
    Minute = sheet.cell(r, 5).value
    WD = sheet.cell(r, 7).value
    WS = sheet.cell(r, 8).value
    WSm = sheet.cell(r, 9).value

    values = (Year, Month, Day, Hour, Minute, WD, WS, WSm)

    try:
        cursor.execute(create_users, values)
    except mysql.connector.errors.DatabaseError:
        print("Cursor", cursor)

cursor.close()
connection.commit()
connection.close()

print("")
print("All Done! Bye, for now.")
print("")

rows = str(sheet.max_row)
print("I just imported 8 columns and ",  rows,  "rows to MySQL!")
